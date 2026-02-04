from openpyxl import Workbook, load_workbook
import os
import Yahoo_sentiment
import Newsapi_sentiment
from datetime import date
today = date.today()

def classify_sentiment(score):
    if score >= 0.15:
        return "Positive"
    elif score <= -0.15:
        return "Negative"
    else:
        return "Neutral"

yahoo_score = Yahoo_sentiment.final_score
newsapi_score = Newsapi_sentiment.final_score

yahoo_ticker = Yahoo_sentiment.ticker
newsapi_query = Newsapi_sentiment.query

yahoo_state = classify_sentiment(yahoo_score)
news_state = classify_sentiment(newsapi_score)

def append_run(ws,run_date,label,score,state,):
    row = 2
    while ws[f"A{row}"].value is not None:
        row += 1
    
    ws[f"A{row}"] = run_date
    ws[f"B{row}"] = label
    ws[f"C{row}"] = score
    ws[f"D{row}"] = state
    
    return row
    


wb = load_workbook("master_file.xlsx")

ws_yahoo = wb["Yahoo Finance Sentiment"]
ws_yahoo['C2'] = float(yahoo_score)
ws_yahoo['B2'] = str(yahoo_ticker)
ws_yahoo['A2'] = today
ws_yahoo['D2'] = yahoo_state
y_row = append_run(ws_yahoo, today, yahoo_ticker, yahoo_score, yahoo_state)

ws_newsapi = wb["News Api Sentiment"]
ws_newsapi['C2'] = float(newsapi_score)
ws_newsapi['B2'] = str(newsapi_query)
ws_newsapi['A2'] = today
ws_newsapi['D2'] = news_state
n_row = append_run(ws_newsapi, today,newsapi_query, newsapi_score, news_state)

wb.save("master_file.xlsx")






